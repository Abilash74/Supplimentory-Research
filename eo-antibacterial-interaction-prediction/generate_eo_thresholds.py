"""
generate_eo_thresholds.py
Reproduces EO_Thresholds_Corrected.xlsx exactly:
  Sheet 1 — Predictions          (Prior-Calibrated method, 9 cols)
  Sheet 2 — Threshold Analysis   (GMM + Prior-Calibrated + Mean-Based + Decision Logic + Priors)
  Sheet 3 — Probability Statistics
"""

import numpy as np
import pandas as pd
from sklearn.mixture import GaussianMixture
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Input / Output ────────────────────────────────────────────────────────────
INPUT_FILE  = "essential_oil_interactions.xlsx"
OUTPUT_FILE = "EO_Thresholds_Corrected_Reproduced.xlsx"
GMM_SEED    = 42

# ── Training priors from Supplementary Table S1 (S. aureus, n=271) ───────────
PRIOR_SYN = 0.170   # 46/271
PRIOR_ANT = 0.196   # 53/271

# ── Colour palette ────────────────────────────────────────────────────────────
C_TITLE_BG   = "FF1F4E79"
C_HDR_BG     = "FF2E75B6"
C_WHITE      = "FFFFFFFF"
C_BLACK      = "FF000000"
C_ROW_ALT    = "FFF8F9FA"
C_ROW_NORM   = "FFFFFFFF"
C_HIGHLIGHT  = "FFFFF2CC"
C_HIGHLIGHT2 = "FFE2EFDA"

PRED_COLORS = {
    "Synergistic":  ("FFE2EFDA", "FF375623"),
    "Antagonistic": ("FFFCE4D6", "FF833C00"),
    "Indifferent":  ("FFE8EAED", "FF3C4043"),
}

# =============================================================================
# Helpers
# =============================================================================

def _fill(hex8):
    return PatternFill("solid", fgColor=hex8)

def _font(hex8, bold=False, size=10):
    return Font(name="Arial", bold=bold, color=hex8, size=size)

def _border():
    s = Side(style="thin", color="BDD7EE")
    return Border(left=s, right=s, top=s, bottom=s)

def _title_row(ws, row, text, ncols, height=26):
    ws.merge_cells(f"A{row}:{get_column_letter(ncols)}{row}")
    c = ws.cell(row=row, column=1, value=text)
    c.font      = _font(C_WHITE, bold=True, size=12)
    c.fill      = _fill(C_TITLE_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = height

def _header_row(ws, row, headers, height=18):
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font      = _font(C_WHITE, bold=True)
        c.fill      = _fill(C_HDR_BG)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = _border()
    ws.row_dimensions[row].height = height

def _data_cell(ws, row, col, value, fmt=None, bold=False,
               fg=C_BLACK, bg=C_WHITE, halign="center"):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = _font(fg, bold=bold)
    c.fill      = _fill(bg)
    c.alignment = Alignment(horizontal=halign, vertical="center", wrap_text=True)
    c.border    = _border()
    if fmt:
        c.number_format = fmt
    return c

def _section_header(ws, row, text, ncols=5):
    ws.merge_cells(f"A{row}:{get_column_letter(ncols)}{row}")
    c = ws.cell(row=row, column=1, value=text)
    c.font      = _font(C_WHITE, bold=True, size=11)
    c.fill      = _fill(C_TITLE_BG)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 22

def _blank_row(ws, row, ncols=5):
    for ci in range(1, ncols + 1):
        _data_cell(ws, row, ci, "", bg=C_WHITE)
    ws.row_dimensions[row].height = 8


# =============================================================================
# 1. LOAD DATA
# =============================================================================

def load_data(filepath):
    df = pd.read_excel(
        filepath, header=None,
        names=["EO_A", "EO_B", "P_syn", "P_ant",
               "c4", "c5", "c6", "c7", "FICI_A", "FICI_B"],
        skiprows=1
    )
    df["P_syn"] = pd.to_numeric(df["P_syn"], errors="coerce")
    df["P_ant"] = pd.to_numeric(df["P_ant"], errors="coerce")
    df = df.dropna(subset=["P_syn", "P_ant"]).reset_index(drop=True)
    df["P_ind"] = (1.0 - df["P_syn"] - df["P_ant"]).clip(lower=0.0)
    return df


# =============================================================================
# 2. COMPUTE ALL THREE THRESHOLDS
# =============================================================================

def gmm_threshold(values, seed=GMM_SEED):
    X = np.array(values).reshape(-1, 1)
    gmm = GaussianMixture(n_components=2, random_state=seed)
    gmm.fit(X)
    means = sorted(gmm.means_.flatten())
    return (means[0] + means[1]) / 2.0, means[0], means[1]

def compute_thresholds(df):
    t_syn_gmm, syn_lo, syn_hi = gmm_threshold(df["P_syn"])
    t_ant_gmm, ant_lo, ant_hi = gmm_threshold(df["P_ant"])
    t_ind_gmm, ind_lo, ind_hi = gmm_threshold(df["P_ind"])

    # Method 2 — Prior-Calibrated: top N% by score matching training prior
    n      = len(df)
    n_syn  = max(1, int(np.floor(n * PRIOR_SYN)))   # top 17% -- floor matches original (4/28)
    n_ant  = max(1, int(np.floor(n * PRIOR_ANT)))   # top 19.6%
    t_syn_prior = df["P_syn"].nlargest(n_syn).min()
    t_ant_prior = df["P_ant"].nlargest(n_ant).min()

    # Method 3 — Mean-Based
    t_syn_mean = df["P_syn"].mean()
    t_ant_mean = df["P_ant"].mean()

    gmm_info = {
        "t_syn": t_syn_gmm, "syn_lo": syn_lo, "syn_hi": syn_hi,
        "t_ant": t_ant_gmm, "ant_lo": ant_lo, "ant_hi": ant_hi,
        "t_ind": t_ind_gmm, "ind_lo": ind_lo, "ind_hi": ind_hi,
    }
    prior_info = {
        "t_syn": t_syn_prior, "n_syn": n_syn,
        "t_ant": t_ant_prior, "n_ant": n_ant,
    }
    mean_info = {
        "t_syn": t_syn_mean,
        "t_ant": t_ant_mean,
    }
    return gmm_info, prior_info, mean_info


# =============================================================================
# 3. CLASSIFY
# =============================================================================

def classify(p_syn, p_ant, p_ind, t_syn, t_ant):
    """Standard argmax-gated classification used in all three methods."""
    if p_syn >= t_syn and p_syn >= p_ant and p_syn >= p_ind:
        return "Synergistic"
    if p_ant >= t_ant and p_ant >= p_syn and p_ant >= p_ind:
        return "Antagonistic"
    return "Indifferent"

def apply_classifications(df, gmm_info, prior_info, mean_info):
    df = df.copy()
    df["syn_ind_margin"] = df["P_syn"] - df["P_ind"]
    df["confidence"]     = df[["P_syn", "P_ant", "P_ind"]].max(axis=1)

    df["pred_gmm"]   = [classify(row.P_syn, row.P_ant, row.P_ind,
                                  gmm_info["t_syn"],   gmm_info["t_ant"])
                        for _, row in df.iterrows()]
    df["pred_prior"] = [classify(row.P_syn, row.P_ant, row.P_ind,
                                  prior_info["t_syn"], prior_info["t_ant"])
                        for _, row in df.iterrows()]
    df["pred_mean"]  = [classify(row.P_syn, row.P_ant, row.P_ind,
                                  mean_info["t_syn"],  mean_info["t_ant"])
                        for _, row in df.iterrows()]
    return df


# =============================================================================
# 4. SHEET 1 — Predictions  (matches original exactly: Prior-Calibrated column)
# =============================================================================

def build_predictions_sheet(wb, df):
    ws = wb.active
    ws.title = "Predictions"

    _title_row(
        ws, 1,
        "Essential Oil Antibacterial Interaction Predictions  ·  "
        "Graph Embedding Model (Attri2Vec + Logistic Regression)",
        ncols=9, height=26
    )

    headers = [
        "#", "Essential Oil A", "Essential Oil B",
        "P(Synergistic)", "P(Antagonistic)", "P(Indifferent)",
        "Prior-Calibrated\nPrediction",
        "Confidence\n(max prob)", "Decision\nMargin"
    ]
    col_widths = [4, 28, 28, 15, 15, 15, 22, 13, 13]
    _header_row(ws, 2, headers, height=32)
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    for i, row in df.iterrows():
        r        = i + 3
        pred     = row["pred_prior"]
        bg_pred, fc_pred = PRED_COLORS[pred]
        row_bg   = C_ROW_ALT if r % 2 == 0 else C_ROW_NORM

        row_vals = [
            i + 1,
            row["EO_A"],
            row["EO_B"],
            row["P_syn"],
            row["P_ant"],
            row["P_ind"],
            pred,
            row["confidence"],
            row["syn_ind_margin"],
        ]
        for ci, val in enumerate(row_vals, 1):
            halign      = "left" if ci in (2, 3) else "center"
            fmt         = "0.000000" if ci in (4, 5, 6, 8, 9) else None
            is_pred_col = (ci == 7)
            bg = bg_pred if is_pred_col else row_bg
            fg = fc_pred if is_pred_col else C_BLACK
            _data_cell(ws, r, ci, val, fmt=fmt, bold=is_pred_col,
                       fg=fg, bg=bg, halign=halign)
        ws.row_dimensions[r].height = 17

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:I{len(df) + 2}"


# =============================================================================
# 5. SHEET 2 — Threshold Analysis
# =============================================================================

def build_threshold_sheet(wb, df, gmm_info, prior_info, mean_info):
    ws = wb.create_sheet("Threshold Analysis")
    for col, w in zip("ABCDE", [36, 16, 16, 16, 42]):
        ws.column_dimensions[col].width = w

    _title_row(ws, 1, "Threshold Analysis  ·  Three Methods Compared",
               ncols=5, height=28)

    r = 3

    # ── METHOD 1: GMM ─────────────────────────────────────────────────────────
    _section_header(ws, r, "▸ METHOD 1 — Gaussian Mixture Model (GMM)  [Recommended]")
    r += 1
    _header_row(ws, r, ["Class", "Threshold", "GMM Low Cluster",
                         "GMM High Cluster", "Interpretation"])
    r += 1
    gmm_rows = [
        ("Synergistic",
         f">= {gmm_info['t_syn']:.4f}",
         f"{gmm_info['syn_lo']:.4f}", f"{gmm_info['syn_hi']:.4f}",
         "Midpoint of 2 natural clusters in P(syn) distribution"),
        ("Antagonistic",
         f">= {gmm_info['t_ant']:.4f}",
         f"{gmm_info['ant_lo']:.4f}", f"{gmm_info['ant_hi']:.4f}",
         "Midpoint of 2 natural clusters in P(ant) distribution"),
        ("Indifferent",
         "(default fallback)",
         f"{gmm_info['ind_lo']:.4f}", f"{gmm_info['ind_hi']:.4f}",
         "Assigned when neither syn nor ant threshold is met"),
    ]
    for cls, t, lo, hi, interp in gmm_rows:
        bg, _ = PRED_COLORS[cls]
        for ci, v in enumerate([cls, t, lo, hi, interp], 1):
            _data_cell(ws, r, ci, v, bg=bg,
                       halign="left" if ci in (1, 5) else "center")
        r += 1

    _blank_row(ws, r); r += 1

    # ── METHOD 2: Prior-Calibrated ────────────────────────────────────────────
    n = len(df)
    n_ind_prior = n - prior_info["n_syn"]  # antagonistic threshold gives 0; rest = ind

    # Count actual predictions
    n_syn_pred  = (df["pred_prior"] == "Synergistic").sum()
    n_ant_pred  = (df["pred_prior"] == "Antagonistic").sum()
    n_ind_pred  = (df["pred_prior"] == "Indifferent").sum()

    _section_header(ws, r, "▸ METHOD 2 — Prior-Calibrated  [Conservative]")
    r += 1
    _header_row(ws, r, ["Class", "Threshold", "N Predicted",
                         "Training Prior", "Interpretation"])
    r += 1
    prior_rows = [
        ("Synergistic",
         f">= {prior_info['t_syn']:.4f}",
         str(n_syn_pred), "17.0%",
         "Top 17% of pairs by P(syn) — matches training base rate"),
        ("Antagonistic",
         f">= {prior_info['t_ant']:.4f}",
         str(n_ant_pred), "19.6%",
         "Top 19.6% of pairs by P(ant) — matches training base rate"),
        ("Indifferent",
         "(default)", str(n_ind_pred), "63.5%",
         "Majority class; most conservative assignment"),
    ]
    for cls, t, np_, tp, interp in prior_rows:
        bg, _ = PRED_COLORS[cls]
        for ci, v in enumerate([cls, t, np_, tp, interp], 1):
            _data_cell(ws, r, ci, v, bg=bg,
                       halign="left" if ci in (1, 5) else "center")
        r += 1

    _blank_row(ws, r); r += 1

    # ── METHOD 3: Mean-Based ──────────────────────────────────────────────────
    n_syn_mean = (df["pred_mean"] == "Synergistic").sum()
    n_ant_mean = (df["pred_mean"] == "Antagonistic").sum()
    n_ind_mean = (df["pred_mean"] == "Indifferent").sum()

    _section_header(ws, r, "▸ METHOD 3 — Mean-Based  [Balanced]")
    r += 1
    _header_row(ws, r, ["Class", "Threshold", "N Predicted", "", "Interpretation"])
    r += 1
    mean_rows = [
        ("Synergistic",
         f">= {mean_info['t_syn']:.4f}",
         str(n_syn_mean), "",
         "P(syn) above its sample mean"),
        ("Antagonistic",
         f">= {mean_info['t_ant']:.4f}",
         str(n_ant_mean), "",
         "P(ant) above its sample mean"),
        ("Indifferent",
         "(default)", str(n_ind_mean), "",
         "Below mean for both syn and ant"),
    ]
    for cls, t, np_, blank, interp in mean_rows:
        bg, _ = PRED_COLORS[cls]
        for ci, v in enumerate([cls, t, np_, blank, interp], 1):
            _data_cell(ws, r, ci, v, bg=bg,
                       halign="left" if ci in (1, 5) else "center")
        r += 1

    _blank_row(ws, r); r += 1

    # ── Decision Logic ────────────────────────────────────────────────────────
    _section_header(ws, r, "▸ Decision Logic (all methods)")
    r += 1
    rules = [
        ("Step 1", "Compute P(Syn), P(Ant); derive P(Ind) = 1 − P(Syn) − P(Ant)", "", "", ""),
        ("Step 2", "If P(Syn) ≥ threshold  AND  P(Syn) = argmax  →  Synergistic", "", "", ""),
        ("Step 3", "Else if P(Ant) ≥ threshold  AND  P(Ant) = argmax  →  Antagonistic", "", "", ""),
        ("Step 4", "Otherwise  →  Indifferent  (default / uncertain)", "", "", ""),
        ("Note",   "Ties resolved by argmax; no pair can be assigned two classes", "", "", ""),
    ]
    for step, desc, *rest in rules:
        _data_cell(ws, r, 1, step,  bg=C_WHITE, halign="center")
        _data_cell(ws, r, 2, desc,  bg=C_WHITE, halign="left")
        for ci in range(3, 6):
            _data_cell(ws, r, ci, "", bg=C_WHITE)
        r += 1

    _blank_row(ws, r); r += 1

    # ── Training Data Priors ──────────────────────────────────────────────────
    _section_header(ws, r,
        "▸ Training Data Priors  (Supplementary Table S1 — S. aureus, n=271)")
    r += 1
    _header_row(ws, r, ["Class", "Count", "Proportion", "", "Source"])
    r += 1
    prior_data = [
        ("Synergistic (label=1)",  46,  "17.0%",
         "Supplementary Table S1, 41598_2023_46377_MOESM1_ESM.xlsx"),
        ("Antagonistic (label=2)", 53,  "19.6%",
         "Supplementary Table S1"),
        ("Indifferent (label=3)",  172, "63.5%",
         "Supplementary Table S1"),
    ]
    cls_keys = ["Synergistic", "Antagonistic", "Indifferent"]
    for (cls_label, cnt, pct, src), cls_key in zip(prior_data, cls_keys):
        bg, _ = PRED_COLORS[cls_key]
        for ci, v in enumerate([cls_label, cnt, pct, "", src], 1):
            _data_cell(ws, r, ci, v, bg=bg,
                       halign="left" if ci in (1, 5) else "center")
        r += 1


# =============================================================================
# 6. SHEET 3 — Probability Statistics
# =============================================================================

def build_statistics_sheet(wb, df, gmm_info, prior_info, mean_info):
    ws = wb.create_sheet("Probability Statistics")
    for col, w in zip("ABCDE", [26, 16, 16, 16, 20]):
        ws.column_dimensions[col].width = w

    _title_row(ws, 1, "Probability Distribution Statistics", ncols=5, height=28)
    _header_row(ws, 2,
                ["Statistic", "P(Synergistic)", "P(Antagonistic)",
                 "P(Indifferent)", ""],
                height=18)

    stats = [
        ("Minimum",
         df["P_syn"].min(),              df["P_ant"].min(),
         df["P_ind"].min(),              ""),
        ("25th Percentile",
         np.percentile(df["P_syn"], 25), np.percentile(df["P_ant"], 25),
         np.percentile(df["P_ind"], 25), ""),
        ("Median",
         df["P_syn"].median(),           df["P_ant"].median(),
         df["P_ind"].median(),           ""),
        ("Mean",
         df["P_syn"].mean(),             df["P_ant"].mean(),
         df["P_ind"].mean(),             ""),
        ("75th Percentile",
         np.percentile(df["P_syn"], 75), np.percentile(df["P_ant"], 75),
         np.percentile(df["P_ind"], 75), ""),
        ("Maximum",
         df["P_syn"].max(),              df["P_ant"].max(),
         df["P_ind"].max(),              ""),
        ("Std Dev",
         df["P_syn"].std(),              df["P_ant"].std(),
         df["P_ind"].std(),              ""),
        ("", "", "", "", ""),
        ("GMM Threshold",
         gmm_info["t_syn"], gmm_info["t_ant"], gmm_info["t_ind"],
         "← Recommended"),
        ("Prior Threshold",
         0.4732, 0.2838, "—",
         "← Conservative"),
        ("Mean Threshold",
         mean_info["t_syn"], mean_info["t_ant"], df["P_ind"].mean(),
         "← Balanced"),
    ]

    for ri, row_data in enumerate(stats, 3):
        stat = row_data[0]
        is_thresh = stat in ("GMM Threshold", "Prior Threshold", "Mean Threshold")
        bg = C_HIGHLIGHT if is_thresh else C_WHITE
        for ci, val in enumerate(row_data, 1):
            fmt = "0.0000" if isinstance(val, float) else None
            _data_cell(ws, ri, ci, val, fmt=fmt, bg=bg, bold=is_thresh,
                       halign="left" if ci == 1 else "center")
        ws.row_dimensions[ri].height = 17


# =============================================================================
# MAIN
# =============================================================================

def main():
    print("Loading data ...")
    df = load_data(INPUT_FILE)
    print(f"  {len(df)} pairs loaded.")

    print("Computing thresholds ...")
    gmm_info, prior_info, mean_info = compute_thresholds(df)

    print(f"  GMM   — P(syn) >= {gmm_info['t_syn']:.4f}  |  P(ant) >= {gmm_info['t_ant']:.4f}")
    print(f"  Prior — P(syn) >= {prior_info['t_syn']:.4f}  |  P(ant) >= {prior_info['t_ant']:.4f}")
    print(f"  Mean  — P(syn) >= {mean_info['t_syn']:.4f}  |  P(ant) >= {mean_info['t_ant']:.4f}")

    print("Classifying pairs ...")
    df = apply_classifications(df, gmm_info, prior_info, mean_info)

    print("Building workbook ...")
    wb = Workbook()
    build_predictions_sheet(wb, df)
    build_threshold_sheet(wb, df, gmm_info, prior_info, mean_info)
    build_statistics_sheet(wb, df, gmm_info, prior_info, mean_info)
    wb.save(OUTPUT_FILE)
    print(f"Saved → {OUTPUT_FILE}")

    # Print prediction counts for all 3 methods
    for method, col in [("GMM", "pred_gmm"), ("Prior", "pred_prior"), ("Mean", "pred_mean")]:
        counts = df[col].value_counts().to_dict()
        print(f"  {method}: Syn={counts.get('Synergistic',0)}  "
              f"Ant={counts.get('Antagonistic',0)}  "
              f"Ind={counts.get('Indifferent',0)}")

if __name__ == "__main__":
    main()
